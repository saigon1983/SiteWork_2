# Класс SKUList представляет из себя набор SKU с методами выборки элементов, сравнения наборов, установки свойств
# для всех элементов набора и сохранения наборов в таблицы Excel
import openpyxl
from openpyxl.utils import get_column_letter as colLetter
from collections import OrderedDict
from Libs.Logic.constants import IGNORE_HEADERS, GROUPS_TREE, USED_HEADERS, PROPERTIES_DATA
from Libs.Logic.class_SKU import SKU

class SKUList:
    BASIC_TABLE_HEADERS = ('Артикул','Тип прибора','Бренд','Модель','Товарная Группа 1','Товарная Группа 2','Товарная Группа 3','Товарная Группа 4','Товарная Группа 5')
    def __init__(self, someList = None):
        # Конструктор принимает в качестве аргумента список или кортеж. Если аргумент явно не передается, создается пустой
        # экземпляр, который в последствии может быть заполнен методом append
        if type(someList)   == tuple:   self.array = list(someList)
        elif type(someList) == list:    self.array = someList[:]
        elif not someList:              self.array = []
        else:                           raise TypeError ('Конструктору класса SKUList передан неверный тип аргумента "{}"'.format(type(someList)))
        self.counter = 0        # Счетчик позиции. Необходим для операций итерации
        self.setupArticleArray()# Список артикулов элементов. Необходим для более быстрого доступа
        self.setupListType()    # Установщик типа набора
# ========== Методы настройки объекта ==========
    def setupArticleArray(self):
        # Метод создает отдельный список артикулов списка SKU для более быстрого поиска элементов по артикулу
        self.articleArray = []
        for sku in self.array:  self.articleArray.append(sku.article)
    def setupListType(self):
        # Метод утановки типа списка. Для гомогенных списков тип равен наименованию ТГ2, для остальных - Смешанный
        self.propertyFields = []
        if self.isHomogenous() and self.firstItem():
            type_1 = self.firstItem().productGroups[1]
            type_2 = self.firstItem().productGroups[2]
            if type_2.lower() in 'подогреватели, вакууматоры панели домино':
                type_2 = ' '.join(X.capitalize() for X in type_2.split(' '))
            if type_1.lower() == 'встраиваемая техника':
                type_2 += ' ВСТР'
            self.type = type_2
            for property in self.firstItem().properties:
                self.propertyFields.append(property)
        else:
            self.type = 'Смешанный'
    def firstItem(self):
        # Внутренний метод, возвращающий первый элемент набора
        try:    return self.array[0]
        except: return None
    def isHomogenous(self):
        # Метод возвращает True, если все его SKU принадлежат одной ТГ2. В таком случае набор считается гомогенным
        if self.array:
            productGroup_1 = self.firstItem().productGroups[1]
            productGroup_2 = self.firstItem().productGroups[2]
            for item in self.array:
                if productGroup_1 != item.productGroups[1] or productGroup_2 != item.productGroups[2]: return False
        return True
# ========== Методы изменения набора ==========
    def append(self, someData):
        # Метод добавления элемента в набор. Элемент обязан быть либо экземпляром класса SKU, либо списком или кортежем
        # экземпляров этого класса. Так же item может быть другим объектом SKUList
        if type(someData) == SKU:
            self.array.append(someData)
            self.articleArray.append(someData.article)
            self.setupListType()
            return
        elif type(someData) == SKUList:
            newData = someData.array[:]
        elif type(someData) == list:
            for item in someData:
                if type(item) != SKU: raise TypeError('Список, переданный в метод append() должен содержать только экземпляры SKU!')
            newData = someData[:]
        elif type(someData) == tuple:
            for item in someData:
                if type(item) != SKU: raise TypeError('Кортеж, переданный в метод append() должен содержать только экземпляры SKU!')
            newData = list(someData)
        else:
            raise TypeError('Аргумент, передаваемый методу append() должен быть типом SKU, SKUList или списокм или кортежем!')
        self.array += newData
        for item in newData:
            self.articleArray.append(item.article)
        self.setupListType()
    def remove(self, item):
        # Метод удаления указанного элемента из набора
        if type(item) != SKU: raise TypeError('Аргумент метода remove() должен быть экземпляром SKU, а получен {}'.format(type(item)))
        try:    self.array.remove(item)
        except: raise ValueError('В наборе нет элемента {}!'.format(item))
        self.setupArticleArray()
        # Производим пересмотр типа набора, если удаленный элемент не соответствовал этому типу
        if item.productGroup[2] != self.type: self.setupListType()
    def removeByArticle(self, article):
        # Метод удаляет из набора элемент с заданным артикулом
        if type(article) != str: raise TypeError('Артикул, передаваемый в метод removeByArticle() должен быть строкой, а получен {}'.format(type(article)))
        try:
            index = self.articleArray.index(article)
            item = self.array[index]
            if item.article == article: self.array.remove(item)
            else: raise ValueError('SKU в позиции {} имеет артикул "{}", отличный от ожидаемого "{}".'.format(index, item.article, article))
        except:
            raise ValueError('Позиции с артикулом {} в наборе нет.'.format(article))
        self.setupArticleArray()
        self.setupListType()
    def removeAllByBrand(self, brand):
        # Метод удаляет из набора все SKU указанного бренда
        needCheckType = False   # Грязный флаг необходимости перепроверки типа набора
        itemsToRemove = []      # Список SKU на удаление
        for item in self.array:
            if item.brand.upper() == brand.upper(): itemsToRemove.append(item)
        for item in itemsToRemove:
            if item.productGroups[2] != self.type: needCheckType = True
            self.array.remove(item)
        self.setupArticleArray()
        if needCheckType: self.setupListType()  # Перепроверяем тип набора при необходимости
    def removeAllByBrands(self, *brands):
        # Метод удаляет из набора все SKU указанных брендов
        needCheckType = False                       # Грязный флаг необходимости перепроверки типа набора
        brands = [brand.upper() for brand in brands]# переводим все бренды в верхний регистр
        itemsToRemove = []                          # Список SKU на удаление
        for item in self.array:
            if item.brand.upper() in brands: itemsToRemove.append(item)
        for item in itemsToRemove:
            if item.productGroups[2] != self.type: needCheckType = True
            self.array.remove(item)
        self.setupArticleArray()
        if needCheckType: self.setupListType()      # Перепроверяем тип набора при необходимости
    def removeDuplicates(self):
        # Метод удаляет все дубликаты из набора
        newArray = []
        for item in self.array:
            if item not in newArray: newArray.append(item)
        self.array = newArray
        self.setupArticleArray()
        self.setupListType()
# ========== Методы получения элементов и выборок ==========
    def splitToHomogenouses(self):
        # Метод разделяет текущий список на несколько гомогенных (по ТГ2) списков и возвращает словарь списков
        resultDict = {}
        resultDict['Прочее'] = SKUList()
        for item in self.array:
            PG = item.productGroups[2]
            if item.productGroups[1] == 'Встраиваемая техника': PG += ' ВСТР'
            if PG:
                if PG not in resultDict.keys():
                    resultDict[PG] = SKUList()
                resultDict[PG].append(item)
            else:
                resultDict['Прочее'].append(item)
        return resultDict
# ========== Методы сохранения набора ==========
    def saveToSimpleTable(self, filename):
        # Метод сохраняет в Excel-таблицу поверхностные данные SKU из набора. Подходит для смешанных наборов
        excelFile = openpyxl.Workbook()
        currentSheet = excelFile.active
        currentSheet.title = 'Список товаров'
        for i in range(len(SKUList.BASIC_TABLE_HEADERS)):
            currentSheet['{}1'.format(openpyxl.utils.get_column_letter(i+1))] = SKUList.BASIC_TABLE_HEADERS[i]
        for item in self.array:
            currentSheet['A{}'.format(self.array.index(item)+2)] = item.article
            currentSheet['B{}'.format(self.array.index(item)+2)] = item.genus
            currentSheet['C{}'.format(self.array.index(item)+2)] = item.brand
            currentSheet['D{}'.format(self.array.index(item)+2)] = item.model
            currentSheet['E{}'.format(self.array.index(item)+2)] = item.productGroups[1]
            currentSheet['F{}'.format(self.array.index(item)+2)] = item.productGroups[2]
            currentSheet['G{}'.format(self.array.index(item)+2)] = item.productGroups[3]
            currentSheet['H{}'.format(self.array.index(item)+2)] = item.productGroups[4]
            currentSheet['I{}'.format(self.array.index(item)+2)] = item.productGroups[5]
        excelFile.save('Output\\' + filename + '.xlsx')
    def saveToDetailedTable(self, filename):
        # Метод создает подробную таблицу со всеми данными о SKU. Работает только с гомогенными наборами
        if not self.isHomogenous(): raise TypeError('Для сохранения в подробную таблицу набор должен быть гомогенным!')
        excelFile = openpyxl.Workbook()
        currentSheet = excelFile.active
        currentSheet.title = self.type
        for i in range(len(SKUList.BASIC_TABLE_HEADERS)):
            currentSheet['{}1'.format(openpyxl.utils.get_column_letter(i+1))] = SKUList.BASIC_TABLE_HEADERS[i]
        i = 0
        for property in self.propertyFields:
            i += 1
            currentSheet['{}1'.format(openpyxl.utils.get_column_letter(len(SKUList.BASIC_TABLE_HEADERS) + i))] = property
        for item in self.array:
            number = self.array.index(item)+2
            currentSheet['A{}'.format(number)] = item.article
            currentSheet['B{}'.format(number)] = item.genus
            currentSheet['C{}'.format(number)] = item.brand
            currentSheet['D{}'.format(number)] = item.model
            currentSheet['E{}'.format(number)] = item.productGroups[1]
            currentSheet['F{}'.format(number)] = item.productGroups[2]
            currentSheet['G{}'.format(number)] = item.productGroups[3]
            currentSheet['H{}'.format(number)] = item.productGroups[4]
            currentSheet['I{}'.format(number)] = item.productGroups[5]
            letter = len(SKUList.BASIC_TABLE_HEADERS)
            for property in self.propertyFields:
                letter += 1
                currentSheet['{}{}'.format(colLetter(letter), number)] = item.properties[property]
        excelFile.save('Output\\' + filename + '.xlsx')

# ========== Перегруженные методы ==========
    def __str__(self):
        # Метод строкового представления объекта списка товаров
        result = ''
        for item in self.array: result += str(item) + '\n'
        return result
    def __len__(self):
        # Перегружаем метод определения длины списка
        return len(self.array)
    def __iter__(self):
        # Метод для реализации итерации списка SKU в цикле for
        return self
    def __next__(self):
        # Метод, определяющий следующий элемент итерации
        if self.counter < len(self.array):
            self.counter += 1
            return self.array[self.counter - 1]
        else:
            self.counter = 0
            raise StopIteration
# ========== Методы класса (внешние конструкторы) ==========
    @classmethod
    def fromExcelTable(cls, givenFile):
        # Метод класса возвращает набор SKU, созданный на основе таблицы Excel, название котрой передается в excelFile
        try:    excelFile = openpyxl.load_workbook(givenFile)
        except: raise FileNotFoundError('Не найден файл "{}". Проверьте наличие файла или путь к нему!'.format(excelFile))
        SHEET = excelFile.worksheets[0] # Выбираем корректную вкладку
        RESULT = []                     # Список результатов
        # Формируем названия ТГ1 и ТГ2 на основе названия вкладки таблицы
        if ' ВСТР' in SHEET.title:
            productGroup_1 = 'Встраиваемая техника'
            productGroup_2 = SHEET.title.rstrip(' ВСТР')
        else:
            for key in GROUPS_TREE.keys():
                for value in GROUPS_TREE[key]:
                    if value == SHEET.title and key != 'Встраиваемая техника':
                        productGroup_1 = key
                        productGroup_2 = value
                        break
        # Формируем список параметров, характерный для данной ТГ2
        key = SHEET.title
        PARAMETERS = []
        for value in PROPERTIES_DATA[key].values():
            if type(value) == str:
                PARAMETERS.append(value)
            else:
                for subvalue in value:
                    PARAMETERS.append(subvalue)
        FIRST_ROW = SHEET[1]            # Получаем список наименований колонок
        # Получаем словарь вида {Наименование поля: индекс колонки}
        HEADERS = {}
        for fieldName in FIRST_ROW:
            if fieldName.value.upper() not in IGNORE_HEADERS:
                HEADERS[fieldName.value] = FIRST_ROW.index(fieldName)
        # Перебираем все строки во вкладке, начиная со второй
        for row in SHEET.iter_rows(min_row = 2):
            # Определяем переменные для данных о товаре
            article = row[HEADERS['Артикул']].value.strip()     # Артикул
            product = row[HEADERS['Название']].value.strip()    # Наименование
            productGroups   = OrderedDict()                     # Товарные группы
            properties      = OrderedDict()                     # Параметры
            productGroups[1] = productGroup_1
            productGroups[2] = productGroup_2
            productGroups[3] = row[HEADERS['Товарная Группа 3']].value
            productGroups[4] = row[HEADERS['Товарная Группа 4']].value
            productGroups[5] = row[HEADERS['Товарная Группа 5']].value
            try:
                for key in PARAMETERS:
                    properties[key] = row[HEADERS[key]].value
            except: print('Неверный ключ "{}" для файла {}|{}'.format(key, productGroup_1, productGroup_2))
            try:    RESULT.append(SKU(product, article, productGroups, properties))
            except: raise KeyError('Проблемы с продуктом "{}"'.format(product))
        return(cls(RESULT))



